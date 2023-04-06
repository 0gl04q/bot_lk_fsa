import datetime
import time
import os
import shutil

import concurrent.futures

from tkinter import messagebox

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException as Sere
from selenium.common.exceptions import ElementClickInterceptedException as E_click
from selenium.common.exceptions import TimeoutException as Te


# ожидание для нажатия на элемент
def wait_click(d, path):
    try:  # Ждем пока элемент не появится, потом на него нажимаем
        WebDriverWait(d, 60).until(
            EC.presence_of_element_located((By.XPATH, path))
        ).click()
    except Sere:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_click(d, path)
    except E_click:
        time.sleep(3)
        wait_click(d, path)
    except Te:
        time.sleep(3)
        wait_click(d, path)


# ожидание для внедрения значения
def wait_send_keys(d, path, paste):
    try:  # Ждем пока элемент не появится, потом вставляем в него значение
        WebDriverWait(d, 60).until(
            EC.presence_of_element_located((By.XPATH, path))
        ).send_keys(paste)
    except Sere:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_send_keys(d, path, paste)
    except E_click:
        time.sleep(3)
        wait_send_keys(d, path, paste)
    except Te:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_send_keys(d, path, paste)


# ожидание появления номера счетчика в поиске
def special_wait(d, path, paste):
    try:
        WebDriverWait(d, 80).until(
            EC.text_to_be_present_in_element((By.XPATH, path), str(paste)))
    except Te:
        d.refresh()

        check_info(d, paste)

        special_wait(d, path, paste)


# поиск по номеру
def check_info(d, paste):
    # Вставка в поле номера рез.пов в соответствующее поле отбора
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div'
                      '/fgis-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis'
                      '-left-panel/div[2]/div[2]/div/div['
                      '1]/fgis-field-input/fgis-field-wrapper/div/div/input', paste)

    time.sleep(5)

    # Нажатие на кнопку найти
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div'
                  '/fgis-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis'
                  '-left-panel/div[3]/div/button')


# Функция создания объекта браузера с необходимыми параметрами
def browse(org):
    options = webdriver.ChromeOptions()

    # Путь и подключение плагина Госуслуг
    path_to_extension = 'crx/gu.crx'
    options.add_extension(path_to_extension)

    # Включаем логи
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    # Отключение метки об автоматическом контроле
    options.add_argument("--disable-software-rasterizer")

    # Добавление пути к профилю Хром
    options.add_argument(fr"user-data-dir=C:\Users\{str(os.getlogin())}\AppData\Local\Google\Chrome\User Data {org}")

    # Путь к драйверу
    path = Service('chromedriver.exe')

    match org:
        case 'МС':

            # Указание папки профиля
            options.add_argument("profile-directory=Profile 1")

        case 'АТМ':

            # Указание папки профиля
            options.add_argument("profile-directory=Profile 2")

            # Путь к драйверу
            path = Service('chromedriver_1.exe')

        case 'СПК':

            # Указание папки профиля
            options.add_argument("profile-directory=Profile 3")

            # Путь к драйверу
            path = Service('chromedriver_2.exe')

    # Создание объекта класса Chrome
    return webdriver.Chrome(service=path, options=options)


# Переход в рабочую область
def work_source(d):
    d.get("http://10.250.74.17/lk")

    wait_click(d, '/html/body/fgis-root/div/fgis-lk/div/fgis-lk-overview/div/div[2]/div/div['
                  '1]/fgis-lk-internal-menu/section/div[3]')

    wait_click(d, '/html/body/fgis-root/div/fgis-activities/div[1]/div/article/div')

    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-attestation/fgis-roei-reestr-selector/div/div[4]')


# Функция внесения сведений о поверках
def update_info(d, r, org):
    # Сообщение для пользователя
    print(f'Заполняю сведения под номером {r[0].value} для организации {org}')

    # Переход в рабочую область
    work_source(d)

    time.sleep(2)  # ожидание

    # Нажатие на кнопку "Добавить"
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div/div["
                  "1]/fgis-table-toolbar/section/div/div[1]/div/fgis-toolbar/div/div[1]/fgis-toolbar-button")

    # Внесение "Номер рез. Поверки СИ"
    wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                      "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                      "/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                      "1]/fgis-card-edit-row[1]/div["
                      "2]/fgis-field-input/fgis-field-wrapper/div/div/input", r[0].value)

    # Внесение "Тип поверяемого средства измерения"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring'
                      '-instruments-card-edit/div/div/div/div/fgis-verification-measuring'
                      '-instruments-card-edit-common/fgis-card-block/div/div['
                      '2]/div/fgis-card-edit-row-two-columns[1]/fgis-card-edit-row[2]/div['
                      '2]/fgis-field-input/fgis-field-wrapper/div/div/input', r[1].value)

    # Внесение "Дата результатов поверки"
    wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                      "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                      "/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                      "2]/fgis-card-edit-row[1]/div["
                      "2]/fgis-field-calendar/fgis-field-wrapper/div/div/fgis-calendar/div/div/inp"
                      "ut", r[2].value)

    # Закрытие формы "Дата результатов поверки" нажатием на экран
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div/div'
                  '/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div[1]/div[1]')

    # Выбор пригодности результата поверки - совершается с помощью 2 действий
    # Открытие списка выбора
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div"
                  "/div/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div["
                  "2]/div/fgis-card-edit-row[1]/div["
                  "2]/fgis-field-selectbox/fgis-field-wrapper/div/div/fgis-selectbox/div/div/div[1]")

    # Выбор действий по условию
    if r[4].value == 'Пригодно':

        # Выбор элемент "Пригодно"
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div["
                      "2]/fgis-virtual-list/div/div[2]/div[1]/fgis-virtual-list-item")

        # Внесение "Дата действия результатов поверки"
        wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments"
                          "-card-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit"
                          "-common/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                          "2]/fgis-card-edit-row[2]/div["
                          "2]/fgis-field-calendar/fgis-field-wrapper/div/div/fgis-calendar/div/div/input", r[3].value)

        # Закрытие формы "Дата результатов поверки" нажатием на экран
        wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div/div'
                      '/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div[1]/div[1]')

    else:
        # Выбор элемент "Непригодно"
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                      "2]/fgis-virtual-list-item")

        # Выбор лица проводившего поверку - совершается с помощью 2 действий
    # Открытие списка выбора
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                  "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                  "/fgis-card-block/div/div[2]/div/fgis-card-edit-row[2]/div["
                  "2]/fgis-field-selectbox/fgis-field-wrapper/div/div/fgis-selectbox/div/div/div[1]")

    time.sleep(5)

    # Внесение в поле фамилии поверителя
    wait_send_keys(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[1]/input", r[5].value)

    time.sleep(3)

    # Выбор фамилии из списка
    if r[5].value == 'Гипич':  # для Гипич отдельное условие так как есть совпадение фамилий
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                      "2]/fgis-virtual-list-item")
    else:
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div["
                      "2]/div/fgis-virtual-list-item")

    # нажатие на подтверждение заполненности формы
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit"
                  "/fgis-verification-measuring-instruments-card-edit-toolbar/div/fgis-toolbar/div/div["
                  "1]/fgis-toolbar-button")

    match org:
        case 'МС':
            paste = 'Общество с ограниченной ответственностью "МС"'
        case 'АТМ':
            paste = 'Общество с ограниченной ответственностью "АТМ"'
        case 'СПК':
            paste = 'Общество с ограниченной ответственностью "СПК"'
        case _:
            raise IOError("Неверно указано имя файла")

    # Подтверждение окончания сохранения черновика
    wait_att(d, paste)

    return 'Черновик РА'


# Ожидание появления необходимого аттрибута
def wait_att(d, paste):
    try:
        WebDriverWait(d, 60).until(
            EC.text_to_be_present_in_element_attribute((By.XPATH, '/html/body/fgis-root/div/fgis-roei'
                                                                  '/fgis-roei-verification-measuring-instruments'
                                                                  '/div/div/div[2]/fgis-table/div[2]/div/table/tbody'
                                                                  '/a[1]/td[10]/div/div[1]/span'), 'title', str(paste)))
    except Te:
        d.refresh()
        wait_att(d, paste)


# Ожидание появления необходимой даты в таблице
def wait_date(d, paste, obsh):
    try:
        match obsh:
            case 1:

                # Принудительное ожидание
                time.sleep(10)

                # Проверка даты в 1 строке таблицы
                WebDriverWait(d, 60).until(
                    EC.text_to_be_present_in_element(
                        (By.XPATH, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification'
                                   '-measuring-instruments/div/div/div[2]/fgis-table/div['
                                   '2]/div/table/tbody/a[1]/td[6]/div/div[1]/span'), paste))

            case _:

                # Проверка первой строки
                WebDriverWait(d, 60).until(
                    EC.text_to_be_present_in_element((By.XPATH, '/html/body/fgis-root/div/fgis-roei/fgis-roei'
                                                                '-verification-measuring-instruments/div/div/div['
                                                                '2]/fgis-table/div[2]/div/table/tbody/a[1]/td['
                                                                '6]/div/div[1]/span'), paste))
                # Проверка второй строки
                WebDriverWait(d, 60).until(
                    EC.text_to_be_present_in_element((By.XPATH, '/html/body/fgis-root/div/fgis-roei/fgis-roei'
                                                                '-verification-measuring-instruments/div/div/div['
                                                                '2]/fgis-table/div[2]/div/table/tbody/a[2]/td['
                                                                '6]/div/div[1]/span'), paste))

        return d.find_element(By.XPATH, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring'
                                        '-instruments/div/div/div[3]/fgis-table-paging/div/div[1]/div['
                                        '2]').text.split()[1]
    except Te:
        wait_date(d, paste, obsh)


# Функция публикации сведения в РА
def public_info(d, num_res):
    # Сообщение для пользователя
    print(f'Публикую сведения {num_res}')

    # Переход в рабочую область
    work_source(d)

    check_info(d, num_res)

    # Ожидание появления строки с необходимым номером
    special_wait(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification'
                    '-measuring-instruments/div/div '
                    '/div[2]/fgis-table/div[2]/div/table/tbody/a[1]/td['
                    '5]/div/div[1]/span', num_res)

    # Выбор первого элемента записи
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div'
                  '/div[2]/fgis-table/div[2]/div/table/tbody/a/td[1]/div')

    # Нажатие на кнопку "Опубликовать"
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div'
                  '/div[1]/fgis-table-toolbar/section/div/div[1]/div/fgis-toolbar/div/div['
                  '4]/fgis-toolbar-button/button')

    # Нажатие на кнопку "Подтверждаю"
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/fgis'
                  '-modal/div/div[1]/div[3]/div/button[1]')

    return 'Опубликовано в РА'


# Функция авторизации в ГУ
def authorization_gu(d, org):
    d.get("https://esia.gosuslugi.ru/login/")

    wait_click(d, '/html/body/esia-root/div/esia-login/div/div[1]/form/div[4]/div/div[2]/div/button')
    wait_click(d, '/html/body/esia-root/div/esia-login/div/div[1]/esia-eds/button')
    match org:
        case 'МС':
            wait_click(d,
                       '/html/body/esia-root/esia-modal/div/div[2]/ng-component/div/esia-eds-item[1]/div')
        case 'АТМ':
            wait_click(d,
                       '/html/body/esia-root/esia-modal/div/div[2]/ng-component/div/esia-eds-item[2]/div')
        case 'СПК':
            wait_click(d,
                       '/html/body/esia-root/esia-modal/div/div[2]/ng-component/div/esia-eds-item[3]/div')

    wait_click(d, '/html/body/esia-root/esia-modal/div/div[2]/ng-component/div/div[2]/a')
    time.sleep(10)


# Функция авторизации в РА
def authorization_ra(d, org):
    # Переход на ФСА после входа в EСИА
    d.get('http://10.250.74.17')

    # ожидание
    # Нажатие на кнопку "Вход в ЕСИА"
    wait_click(d, "/html/body/div/div[2]/div/div[2]/a[1]")

    # Переключение рабочей вкладки
    windows = d.window_handles  # Список всех вкладок

    time.sleep(5)  # ожидание

    d.close()  # Закрываем первое окно
    d.switch_to.window(windows[1])  # Переключаем драйвер на рабочую область

    time.sleep(5)  # ожидание

    match org:
        case 'МС':
            # Выбор организации "МС"
            wait_click(d, "/html/body/div/div[2]/div/div/table/tbody/tr[2]/td/a")

            # Открытие списка выбора рабочей области
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div["
                          "2]/fgis-selectbox/div/div/div[2]")

            # Выбор рабочей области в нашем случае АЛ
            wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                          "1]/fgis-virtual-list-item")

            # Ожидание
            time.sleep(5)

            # Нажатие на кнопку входа в личный кабинет
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div[2]/button")

        case 'АТМ':
            # Выбор организации "АТМ"
            wait_click(d, '/html/body/div/div[2]/div/div/table/tbody/tr[2]/td/a')

            # Открытие списка выбора рабочей области
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div["
                          "2]/fgis-selectbox/div/div/div[2]")

            # Выбор рабочей области в нашем случае АЛ
            wait_click(d,
                       '/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div['
                       '1]/fgis-virtual-list-item/li/div')

            # Ожидание
            time.sleep(5)

            # Нажатие на кнопку входа в личный кабинет
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div[2]/button")

        case 'СПК':
            # Выбор организации "СПК"
            wait_click(d, '/html/body/div/div[2]/div/div/table/tbody/tr/td/a')

    # Ожидание
    time.sleep(5)


# Функция для определения продолжить работу или начать авторизацию
def all_authorization(d, org):
    # переход в рабочую область
    d.get('http://10.250.74.17/roei/verification-measuring-instruments')

    # Ожидание появления строки с необходимым номером
    try:
        WebDriverWait(d, 10).until(EC.text_to_be_present_in_element((By.XPATH, '/html/body/fgis-root/div/fgis-roei'
                                                                               '/fgis-roei-verification-measuring'
                                                                               '-instruments/div/fgis-roei'
                                                                               '-verification-measuring-instruments'
                                                                               '-advanced-search/fgis-filters-panel'
                                                                               '/fgis-left-panel/div[3]/div/a'),
                                                                    'Очистить фильтры'))
    except Te:

        # Сообщение пользователю
        print(f'Начинаю прохождение авторизации в ГУ для организации {org}')

        # прохождение авторизации В ГУ
        authorization_gu(d, org)

        # Сообщение пользователю
        print(f'Закончил прохождение авторизации в ГУ для организации {org}')

        # ожидание
        time.sleep(5)

        # Сообщение пользователю
        print(f'Начинаю прохождение авторизации в РА для организации {org}')

        # прохождение авторизации в RA
        authorization_ra(d, org)

        # Сообщение пользователю
        print(f'Закончил прохождение авторизации в РА для организации {org}')


# Функция сохранения информации в файле Excel
def save_wb(wb, file):
    try:
        wb.save(file)  # сохраняем статус

    except PermissionError:  # ожидание доступности файла

        # Сообщение для пользователя
        print(f'Нет доступа к файлу {file}')

        # Ожидание + перезагрузка
        time.sleep(5)
        save_wb(wb, file)


# Функция проверки количества черновиков
def check_kol(d, my_date, num, prom, file):
    # Переход в рабочую область
    work_source(d)

    time.sleep(2)  # ожидание

    # Внесение с "Дата поверки"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis'
                      '-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left'
                      '-panel/div[2]/div[2]/div/div[2]/fgis-date-range/div/fgis-calendar[1]/div/div/input',
                   my_date)

    # Внесение по "Дата поверки"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis'
                      '-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left'
                      '-panel/div[2]/div[2]/div/div[2]/fgis-date-range/div/fgis-calendar[2]/div/div/input',
                   my_date)

    # Включить отображение кол-ва поверок
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis-roei'
                  '-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left-panel/div['
                  '2]/div[2]/label')

    # Нажимаем кнопку найти
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis-roei'
                  '-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left-panel/div['
                  '3]/div/button')

    # Проверка даты в 2-х первых полях таблицы и получение итогового значения
    count_poverka_ra = wait_date(d, my_date, num)

    # Проверка на совпадение
    if not count_poverka_ra == str(num):
        messagebox.showwarning(f"Остановка программы {file}",
                               f'Не сходятся значения кол-ва счетчиков\n'
                               f'Количество в РА - {count_poverka_ra}\n'
                               f'Количество черновиков в файле - {prom}\n'
                               f'Общее количество поверок в файле - {num} \n')

        print('Остановка программы в связи с расхождением кол-ва, ожидание действий пользователя')

        time.sleep(100000)


# Функция скачивания файла из РА и проверки количества опубликованных счетчиков
def download_file(d, my_date, obsh, prom, file):

    # Сообщение для пользователя
    print(f'Начинаю проверку публикации файла {file}')

    # Переход в рабочую область
    work_source(d)

    time.sleep(2)  # ожидание

    # Внесение с "Дата поверки"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis'
                      '-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left'
                      '-panel/div[2]/div[2]/div/div[2]/fgis-date-range/div/fgis-calendar[1]/div/div/input',
                   my_date)

    # Внесение по "Дата поверки"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis'
                      '-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left'
                      '-panel/div[2]/div[2]/div/div[2]/fgis-date-range/div/fgis-calendar[2]/div/div/input',
                   my_date)

    # Включить отображение кол-ва поверок
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis-roei'
                  '-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left-panel/div['
                  '2]/div[2]/label')

    # Нажимаем кнопку найти
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis-roei'
                  '-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left-panel/div['
                  '3]/div/button')

    # Проверка даты в 2-х первых полях таблицы и получение итогового значения
    count_poverka_ra = wait_date(d, my_date, obsh)

    # Проверка на совпадение
    if not count_poverka_ra == str(obsh):
        messagebox.showwarning(f"Остановка программы {file}",
                               f'Не сходятся значения кол-ва счетчиков\n'
                               f'Количество в РА - {count_poverka_ra}\n'
                               f'Количество опубликованных в файле - {prom}\n'
                               f'Общее количество поверок в файле - {obsh} \n')

        print('Остановка программы в связи с расхождением кол-ва, ожидание действий пользователя')

        time.sleep(100000)

    # Сообщение пользователю
    print(f'Проверка файла {file} пройдена успешно, приступаю к выгрузке файла из РА')

    # Нажатие кнопки выгрузки
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div/div['
                  '1]/fgis-table-toolbar/section/div/div[2]/fgis-toolbar/div/div[1]/fgis-toolbar-button/button')

    # Выбор пункта excel
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/fgis-export'
                  '-settings-modal/fgis-modal/div/div[1]/div[2]/div[2]/p-radiobutton[2]/label')

    # Нажатие на кнопку Экспорт
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/fgis-export'
                  '-settings-modal/fgis-modal/div/div[1]/div[3]/div/fgis-export-step/button[1]')

    # Ожидание скачивания файла
    wait_clickable(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/fgis-roei'
                      '-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis-left-panel/div['
                      '2]/div[2]/div/div[1]/fgis-field-input/fgis-field-wrapper/div/div/input')

    # Имя файла
    name_file = f'Сведения по обеспечению ' \
                f'единства измерений (Поверка СИ) от {str(datetime.datetime.today().strftime("%d.%m.%Y"))}.xlsx'

    # путь к скачанному файлу
    path_to_file = f'C:/Users/{str(os.getlogin())}/Downloads/{name_file}'

    # Ожидание появления файла
    while not os.path.exists(path_to_file):
        time.sleep(5)

    # путь итог
    new_path = f'end_ra_excel/{name_file.replace(".xlsx", f" - {file}")}'

    # Перемещение законченного файла
    shutil.move(path_to_file, new_path)

    # Сообщение пользователю
    print(f'Файл {new_path.split("/")[1]} выгружен успешно !!!')


# ожидание появления номера счетчика в поиске
def wait_clickable(d, path):
    try:
        WebDriverWait(d, 60).until(
            EC.element_to_be_clickable((By.XPATH, path)))
    except Te:
        wait_clickable(d, path)


# Функция обработки файла
def one_rm(f_name, organization):
    chern = 0
    publ_ra = 0
    obsh = 0

    # Создание объекта класса Chrome
    driver = browse(organization)

    # Проверка на необходимость авторизации
    all_authorization(driver, organization)

    file_name = f'file/{f_name}'

    # Сообщение для пользователя
    print(f'Начинаю перебор файла {f_name}')

    # Открытие файла Excel
    wb = load_workbook(filename=file_name)
    sheet = wb.active

    # Сообщение для пользователя
    print(f'Приступаю к созданию черновиков для организации {organization}')

    # Проходим по строкам
    for row in sheet.iter_rows(min_row=2):

        # проверка на обязательное наличие сведений
        if row[1].value:

            # проверка поля статуса Если статус Выгружена в АРШИН тогда переходим к созданию черновика
            if row[6].value == 'Выгружена в АРШИН':
                row[6].value = update_info(driver, row, organization)  # Заполняем и сохраняем черновик
                save_wb(wb, file_name)

                # Сообщение для пользователя
                print(f'Черновик {row[0].value} сохранен для организации {organization}')

            # Подсчет счетчиков со статусом Черновик РА
            if row[6].value == 'Черновик РА':
                chern += 1

            # сбор общего кол-ва поверок
            obsh += 1

        else:
            break

    # Сообщение для пользователя
    print(f'Все сведения загружены, приступаю к проверке кол-ва черновиков для организации {organization}')

    my_date = '.'.join(f_name.split()[1].split('.')[0:3])

    # Проверка количества счетчиков
    check_kol(driver, my_date, obsh, chern, f_name)

    # Сообщение для пользователя
    print(f'Проверка прошла успешно, приступаю к публикации черновиков для организации {organization}')

    # Проходим по строкам
    for row in sheet.iter_rows(min_row=2):

        # проверка на обязательное наличие сведений
        if row[1].value:

            # Если статус = Черновик РА тогда приступаем к публикации сведения
            if row[6].value == 'Черновик РА':
                time.sleep(3)
                row[6].value = public_info(driver, row[0].value)  # публикуем запись

                time.sleep(3)
                save_wb(wb, file_name)  # сохраняем статус

                # Сообщение для пользователя
                print(f'Сведение под номером {row[0].value} опубликованы для организации {organization}')

            # Подсчет опубликованных поверок
            if row[6].value == 'Опубликовано в РА':
                publ_ra += 1

    # Закрытие рабочей книга
    wb.close()

    # Перемещение законченного файла
    shutil.move(file_name, f'end/{f_name}')

    # Ожидание
    time.sleep(5)

    # Скачивание и перенос файла в необходимую папку
    download_file(driver, my_date, obsh, publ_ra, f_name)

    # Закрытие браузера
    driver.close()

    # создание файла итогов
    print_end_publ(publ_ra, file_name)


# Добавление итогов по загрузке в файл
def print_end_publ(p, file):
    with open('end/log.txt', 'a', encoding='utf-8') as write_file:
        name = file.split("/")[1].replace(".xlsx", "").split('.')
        write_file.write(f'{name[0]}.{name[1]} - {p}\n')

        # Сообщение пользователю
        print(f'Файл {file.split("/")[1]} завершен. Записал сведения о выгруженных поверках в файл log.txt')


# Функция проверки папок и необходимых файлов
def check_path():
    if not os.path.exists('end'):
        os.mkdir('end')
    if not os.path.exists('file'):
        os.mkdir('file')
    if not os.path.exists('end_ra_excel'):
        os.mkdir('end_ra_excel')


def thread_function(tup):
    print(f'Запуск потока {tup[0]}')

    # Перебор файлов организации
    for book in tup[1]:
        one_rm(book, tup[0])

    print(f'Конец потока {tup[0]}')


# Основная функция
if __name__ == '__main__':

    # проверка пути
    check_path()

    # Получение списка файлов в папке
    books = os.listdir(f'{os.getcwd()}/file')

    # Создание списков по организациям
    atm_list = list(filter(lambda x: x.split()[0] == 'АТМ', books))
    ms_list = list(filter(lambda x: x.split()[0] == 'МС', books))
    spk_list = list(filter(lambda x: x.split()[0] == 'СПК', books))

    # Запуск потока как отдельного переборщика для каждой организации
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        executor.map(thread_function, [('АТМ', atm_list), ('МС', ms_list), ('СПК', spk_list)])
