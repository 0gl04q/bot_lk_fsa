import time
import os
import shutil

from openpyxl import load_workbook
from threading import Thread

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException as Sere
from selenium.common.exceptions import ElementClickInterceptedException as E_click
from selenium.common.exceptions import TimeoutException as Te


# ожидание для нажатия на элемент
def wait_click(d, path):
    try:  # Ждем пока элемент не появится, потом на него нажимаем
        WebDriverWait(d, 60).until(
            EC.presence_of_element_located((By.XPATH, path))
        ).click()
    except Sere:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_click(d, path)
    except E_click:
        time.sleep(3)
        wait_click(d, path)
    except Te:
        time.sleep(3)
        wait_click(d, path)


# ожидание для внедрения значения
def wait_send_keys(d, path, paste):
    try:  # Ждем пока элемент не появится, потом вставляем в него значение
        WebDriverWait(d, 60).until(
            EC.presence_of_element_located((By.XPATH, path))
        ).send_keys(paste)
    except Sere:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_send_keys(d, path, paste)
    except E_click:
        time.sleep(3)
        wait_send_keys(d, path, paste)
    except Te:  # обрабатываем исключение ожиданием и рекурсией
        time.sleep(3)
        wait_send_keys(d, path, paste)


# ожидание появления номера счетчика в поиске
def special_wait(d, path, paste):
    try:
        WebDriverWait(d, 80).until(
            EC.text_to_be_present_in_element((By.XPATH, path), str(paste)))
    except Te:
        d.refresh()

        check_info(d, paste)

        special_wait(d, path, paste)


# поиск по номеру
def check_info(d, paste):
    # Вставка в поле номера рез.пов в соответствующее поле отбора
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div'
                      '/fgis-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis'
                      '-left-panel/div[2]/div[2]/div/div['
                      '1]/fgis-field-input/fgis-field-wrapper/div/div/input', paste)

    time.sleep(5)

    # Нажатие на кнопку найти
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div'
                  '/fgis-roei-verification-measuring-instruments-advanced-search/fgis-filters-panel/fgis'
                  '-left-panel/div[3]/div/button')


# Функция создания объекта браузера с необходимыми параметрами
def browse(org):
    options = webdriver.ChromeOptions()

    # Путь и подключение плагина Госуслуг
    path_to_extension = 'crx/gu.crx'
    options.add_extension(path_to_extension)

    # Отключение метки об автоматическом контроле
    options.add_argument("--disable-blink-features=AutomationControlled")

    match org:
        case 'МС':
            # Добавление пути к профилю Хром
            options.add_argument(fr"user-data-dir=C:\Users\{str(os.getlogin())}\AppData\Local\Google\Chrome\User Data")

            # Указание папки профиля
            options.add_argument("profile-directory=Profile 1")

            # Путь к драйверу
            path = Service('chromedriver_1.exe')

            # Создание объекта класса Chrome
            return webdriver.Chrome(service=path, options=options)

        case 'АТМ':
            # Добавление пути к профилю Хром
            options.add_argument(
                fr"user-data-dir=C:\Users\{str(os.getlogin())}\AppData\Local\Google\Chrome\User Data Copy")

            # Указание папки профиля
            options.add_argument("profile-directory=Profile 2")

            # Путь к драйверу
            path = Service('chromedriver.exe')

            # Создание объекта класса Chrome
            return webdriver.Chrome(service=path, options=options)


# Переход в рабочую область
def work_source(d):
    d.get("http://10.250.74.17/lk")

    wait_click(d, '/html/body/fgis-root/div/fgis-lk/div/fgis-lk-overview/div/div[2]/div/div['
                  '1]/fgis-lk-internal-menu/section/div[3]')

    wait_click(d, '/html/body/fgis-root/div/fgis-activities/div[1]/div/article/div')

    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-attestation/fgis-roei-reestr-selector/div/div[4]')


# Функция внесения сведений о поверках
def update_info(d, r, org):

    # Переход в рабочую область
    work_source(d)

    time.sleep(2)  # ожидание

    # Нажатие на кнопку "Добавить"
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div/div["
                  "1]/fgis-table-toolbar/section/div/div[1]/div/fgis-toolbar/div/div[1]/fgis-toolbar-button")

    # Внесение "Номер рез. поверки СИ"
    wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                      "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                      "/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                      "1]/fgis-card-edit-row[1]/div["
                      "2]/fgis-field-input/fgis-field-wrapper/div/div/input", r[0].value)

    # Внесение "Тип поверяемого средства измерения"
    wait_send_keys(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring'
                      '-instruments-card-edit/div/div/div/div/fgis-verification-measuring'
                      '-instruments-card-edit-common/fgis-card-block/div/div['
                      '2]/div/fgis-card-edit-row-two-columns[1]/fgis-card-edit-row[2]/div['
                      '2]/fgis-field-input/fgis-field-wrapper/div/div/input', r[1].value)

    # Внесение "Дата результатов поверки"
    wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                      "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                      "/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                      "2]/fgis-card-edit-row[1]/div["
                      "2]/fgis-field-calendar/fgis-field-wrapper/div/div/fgis-calendar/div/div/inp"
                      "ut", r[2].value)

    # Закрытие формы "Дата результатов поверки" нажатием на экран
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div/div'
                  '/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div[1]/div[1]')

    # Выбор пригодности результата поверки - совершается с помощью 2 действий
    # Открытие списка выбора
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div"
                  "/div/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div["
                  "2]/div/fgis-card-edit-row[1]/div["
                  "2]/fgis-field-selectbox/fgis-field-wrapper/div/div/fgis-selectbox/div/div/div[1]")

    # Выбор действий по условию
    if r[4].value == 'Пригодно':

        # Выбор элемент "Пригодно"
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div["
                      "2]/fgis-virtual-list/div/div[2]/div[1]/fgis-virtual-list-item")

        # Внесение "Дата действия результатов поверки"
        wait_send_keys(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments"
                          "-card-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit"
                          "-common/fgis-card-block/div/div[2]/div/fgis-card-edit-row-two-columns["
                          "2]/fgis-card-edit-row[2]/div["
                          "2]/fgis-field-calendar/fgis-field-wrapper/div/div/fgis-calendar/div/div/input", r[3].value)

        # Закрытие формы "Дата результатов поверки" нажатием на экран
        wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit/div/div/div'
                      '/div/fgis-verification-measuring-instruments-card-edit-common/fgis-card-block/div/div[1]/div[1]')

    else:
        # Выбор элемент "Непригодно"
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                      "2]/fgis-virtual-list-item")

        # Выбор лица проводившего поверку - совершается с помощью 2 действий
    # Открытие списка выбора
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card"
                  "-edit/div/div/div/div/fgis-verification-measuring-instruments-card-edit-common"
                  "/fgis-card-block/div/div[2]/div/fgis-card-edit-row[2]/div["
                  "2]/fgis-field-selectbox/fgis-field-wrapper/div/div/fgis-selectbox/div/div/div[1]")

    time.sleep(5)

    # Внесение в поле фамилии поверителя
    wait_send_keys(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[1]/input", r[5].value)

    time.sleep(3)

    # Выбор фамилии из списка
    if r[5].value == 'Гипич':  # для Гипич отдельное условие так как есть совпадение фамилий
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                      "2]/fgis-virtual-list-item")
    else:
        wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div["
                      "2]/div/fgis-virtual-list-item")

    # нажатие на подтверждение заполненности формы
    wait_click(d, "/html/body/fgis-root/div/fgis-roei/fgis-verification-measuring-instruments-card-edit"
                  "/fgis-verification-measuring-instruments-card-edit-toolbar/div/fgis-toolbar/div/div["
                  "1]/fgis-toolbar-button")



    match org:
        case 'МС':
            paste = 'Общество с ограниченной ответственностью "МС"'
        case 'АТМ':
            paste = 'Общество с ограниченной ответственностью "АТМ"'
        case _:
            raise IOError("Неверно указано имя файла")

    wait_att(d, paste)

    return 'Черновик РА'


# Ожидание появления необходимого аттрибута
def wait_att(d, paste):
    try:
        WebDriverWait(d, 60).until(
            EC.text_to_be_present_in_element_attribute((By.XPATH, '/html/body/fgis-root/div/fgis-roei'
                                                                  '/fgis-roei-verification-measuring-instruments'
                                                                  '/div/div/div[2]/fgis-table/div[2]/div/table/tbody'
                                                                  '/a[1]/td[10]/div/div[1]/span'), 'title', str(paste)))
    except Te:
        d.refresh()
        wait_att(d, paste)


# Функция публикации сведения в РА
def public_info(d, num_res):

    # Переход в рабочую область
    work_source(d)

    check_info(d, num_res)

    # Ожидание появления строки с необходимым номером
    special_wait(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification'
                    '-measuring-instruments/div/div '
                    '/div[2]/fgis-table/div[2]/div/table/tbody/a[1]/td['
                    '5]/div/div[1]/span', num_res)

    # Выбор первого элемента записи
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div'
                  '/div[2]/fgis-table/div[2]/div/table/tbody/a/td[1]/div')

    # Нажатие на кнопку "Опубликовать"
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/div/div'
                  '/div[1]/fgis-table-toolbar/section/div/div[1]/div/fgis-toolbar/div/div['
                  '4]/fgis-toolbar-button/button')

    # Нажатие на кнопку "Подтверждаю"
    wait_click(d, '/html/body/fgis-root/div/fgis-roei/fgis-roei-verification-measuring-instruments/fgis'
                  '-modal/div/div[1]/div[3]/div/button[1]')

    return 'Опубликовано в РА'


# Функция авторизации в ГУ
def authorization_gu(d, org):
    d.get("https://esia.gosuslugi.ru/login/")

    wait_click(d, '/html/body/esia-root/div/esia-login/div/div[1]/form/div[4]/div/div[2]/div/button')
    wait_click(d, '/html/body/esia-root/div/esia-login/div/div[1]/esia-eds/button')
    match org:
        case 'МС':
            wait_click(d,
                       '/html/body/esia-root/esia-modal-placeholder/div/div[2]/div/div/div/ng-component/div/esia-eds-item[1]')
        case 'АТМ':
            wait_click(d,
                       '/html/body/esia-root/esia-modal-placeholder/div/div[2]/div/div/div/ng-component/div/esia-eds-item[2]')

    wait_click(d, '/html/body/esia-root/esia-modal-placeholder/div/div[2]/div/div/div/ng-component/div/div[2]/a')
    time.sleep(10)


# Функция авторизации в РА
def authorization_ra(d, org):
    # Переход на ФСА после входа в EСИА
    work_source(d)

    # ожидание
    # Нажатие на кнопку "Вход в ЕСИА"
    wait_click(d, "/html/body/div/div[2]/div/div[2]/a[1]")

    # Переключение рабочей вкладки
    windows = d.window_handles  # Список всех вкладок

    time.sleep(5)  # ожидание

    d.close()  # Закрываем первое окно
    d.switch_to.window(windows[1])  # Переключаем драйвер на рабочую область

    time.sleep(5)  # ожидание

    match org:
        case 'МС':
            # Выбор организации "МС"
            wait_click(d, "/html/body/div/div[2]/div/div/table/tbody/tr[2]/td/a")

            # Открытие списка выбора рабочей области
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div["
                          "2]/fgis-selectbox/div/div/div[2]")

            # Выбор рабочей области в нашем случае АЛ
            wait_click(d, "/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div["
                          "1]/fgis-virtual-list-item")

        case 'АТМ':
            # Выбор организации "АТМ"
            wait_click(d, '/html/body/div/div[2]/div/div/table/tbody/tr[2]/td/a')

            # Открытие списка выбора рабочей области
            wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div["
                          "2]/fgis-selectbox/div/div/div[2]")

            # Выбор рабочей области в нашем случае АЛ
            wait_click(d,
                       '/html/body/fgis-root/fgis-select-dropdown/div/div/div[2]/fgis-virtual-list/div/div[2]/div[1]/fgis-virtual-list-item/li/div')

    # Ожидание
    time.sleep(5)

    # Нажатие на кнопку входа в личный кабинет
    wait_click(d, "/html/body/fgis-root/div/fgis-lk/div/fgis-lk-selector/div[5]/div/div[2]/button")

    # Ожидание
    time.sleep(5)


# Функция для определения продолжить работу или начать авторизацию
def all_authorization(d, org):
    # переход в рабочую область
    work_source(d)

    # Ожидание появления строки с необходимым номером
    try:
        WebDriverWait(d, 10).until(EC.text_to_be_present_in_element((By.XPATH, '/html/body/fgis-root/div/fgis'
                                                                               '-header/header/div['
                                                                               '1]/fgis-title/div/div'),
                                                                    'Сведения по обеспечению единства '
                                                                    'измерений'))
    except Te:
        # прохождение авторизации В ГУ
        authorization_gu(d, org)

        # ожидание
        time.sleep(5)

        # прохождение авторизации в RA
        authorization_ra(d, org)


# Функция обработки файла
def one_rm(file_name, organization):

    publ_ra = 0

    # Создание объекта класса Chrome
    driver = browse(organization)

    # Проверка на необходимость авторизации
    all_authorization(driver, organization)

    # Открытие файла Excel
    wb = load_workbook(filename=file_name)
    sheet = wb.active

    # Проходим по строкам
    for row in sheet.iter_rows(min_row=2):

        # проверка на обязательное наличие сведений
        if row[1].value:

            # проверка поля статуса Если статус Выгружена в АРШИН тогда переходим к созданию черновика
            if row[6].value == 'Выгружена в АРШИН':
                row[6].value = update_info(driver, row, organization)  # Заполняем и сохраняем черновик
                wb.save(file_name)  # сохраняем статус

            # Если статус = Черновик РА тогда приступаем к публикации сведения
            if row[6].value == 'Черновик РА':
                time.sleep(3)
                row[6].value = public_info(driver, row[0].value)  # публикуем запись\
                time.sleep(3)
                wb.save(file_name)  # сохраняем статус

            if row[6].value == 'Опубликовано в РА':
                publ_ra += 1
        else:
            break

    # Закрытие рабочей книга
    wb.close()

    # Перемещение законченного файла
    shutil.move(file_name, f'end/{file_name.split("/")[1]}')

    # Ожидание
    time.sleep(5)

    # Закрытие браузера
    driver.close()

    # создание файла итогов
    print_end_publ(publ_ra, file_name)


# Добавление итогов по загрузке в файл
def print_end_publ(p, file):
    with open('end/log.txt', 'a', encoding='utf-8') as write_file:
        name = file.split("/")[1].replace(".xlsx", "").split('.')
        write_file.write(f'{name[0]}.{name[1]} - {p}\n')


# Функция создания потока
def add_tread(excel):
    # Запуск процесса
    tr = Thread(name='Browser for {}'.format(excel), target=one_rm, args=(f'file/{excel}', excel.split()[0],))
    tr.start()

    print(tr.name + ' started!')
    return tr


# Функция закрытия потоков
def close_tread(tr_list):

    # Ожидание завершения процессов
    for tr in tr_list:
        tr.join()


# Функция проверки папок и необходимых файлов
def check_path():
    if not os.path.exists('end'):
        os.mkdir('end')
    if not os.path.exists('file'):
        os.mkdir('file')


# Основная функция
if __name__ == '__main__':

    # проверка пути
    check_path()

    # Получение списка файлов в папке
    books = os.listdir(f'{os.getcwd()}/file')

    # Создание списков по организациям
    atm_list = list(filter(lambda x: x.split()[0] == 'АТМ', books))
    ms_list = list(filter(lambda x: x.split()[0] == 'МС', books))

    # Формирование списка кортежей
    sp_tuple = list(zip(atm_list, ms_list))

    thread_list = []

    # Цикл прохода по списку кортежей
    for b in sp_tuple:
        book_atm, book_mc = b

        # Запуск процесса АТМ
        thread_list.append(add_tread(book_atm))

        # Ожидание перед запуском второго браузера
        time.sleep(10)

        # Запуск процесса МС
        thread_list.append(add_tread(book_mc))

        # Ожидание завершения процессов
        close_tread(thread_list)

    # Поиск оставшихся файлов в папке
    # Проверяем отличие списка кортежей от суммы списком одной и второй организации
    if len(sp_tuple) * 2 != len(atm_list) + len(ms_list):

        # Создаем список из кортежей
        list_book = [b for t_books in sp_tuple for b in t_books]

        # Создаем список
        itg = filter(lambda x: x not in list_book, books)

        # Перебираем список остатков
        for book in itg:
            # Запускаем поток
            thread_list.append(add_tread(book))

            # Ожидание завершения процессов
            close_tread(thread_list)
